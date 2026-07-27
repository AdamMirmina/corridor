"""
Executive-director timeline sheet, matching the format of a BID executive-
director timeline reference: one row per organization, one column per year
from the earliest founding year among the sampled orgs through the present.
"N/A" before an org existed, "Full Name (Initials)" on the year a director
started, bare initials for each year they continued, blank where the org
existed but no leader is confirmed for that year.

The 20-organization sample is drawn only from orgs with reliable data (a
matched EIN, 5+ years of digitized Form 990 filings, and at least one named
leader) so the timeline isn't mostly blank rows, and it's frozen in
data/ed_timeline_sample.txt so it doesn't reshuffle every time this is
re-run; regenerate that file deliberately (see the random.seed in the git
history) if a new sample is ever wanted.

Sources, in priority order, per year:
  1. Direct research (output/leadership_history.csv, parsed tenure ranges)
  2. Corridor's own current-executive match (web/src/data/dataset.json),
     used only to extend a name through to the present when that research
     ends before now and the pipeline's own EIN-matched executive is
     available.

Run after tools/build_outputs.py (needs its output/corridor_dataset.xlsx and
web/src/data/dataset.json). Appends an "ED Timeline (sample)" sheet to the
existing workbook.

Out: adds a sheet to output/corridor_dataset.xlsx (then re-copied to
web/public by re-running tools/build_outputs.py's copy_downloads, or copy
manually).
"""

import csv
import json
import os
import re

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(HERE, "data")
OUT = os.path.join(HERE, "output")
WEB_DATA = os.path.join(HERE, "web", "src", "data")


def norm(name):
    name = name.lower()
    name = re.sub(r"[^a-z0-9 ]", " ", name)
    for w in ["the", "inc", "incorporated", "corporation", "corp", "cdc", "community",
              "development", "association", "company", "companies", "group", "of", "and"]:
        name = re.sub(rf"\b{w}\b", " ", name)
    return re.sub(r"\s+", " ", name).strip()


def initials(person):
    # "Jamila Harris-Morrison - Executive Director" -> "Jamila Harris-Morrison", "JHM"
    name = person.split(" - ")[0].strip()
    words = [w for w in re.split(r"[\s-]+", name) if w and w[0].isalpha()]
    ini = "".join(w[0].upper() for w in words[:3])
    return name, ini


def load_leadership_history():
    by_org = {}
    with open(os.path.join(OUT, "leadership_history.csv"), newline="", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            by_org.setdefault(norm(r["org_name"]), []).append(r)
    return by_org


def load_founding_years():
    """Founding year per org, preferring Ben's own 'years_active' research
    (e.g. "1934-2000s", "1985-present") over the leadership_history's
    earliest start_year, since some orgs (e.g. Germantown Settlement) have a
    known founding year but no confirmed named leader from that far back."""
    founding = {}
    with open(os.path.join(DATA, "ben_research.json"), encoding="utf-8") as f:
        ben_orgs = json.load(f)
    for o in ben_orgs:
        m = re.search(r"\b(18|19|20)\d{2}\b", o.get("years_active") or "")
        if m:
            founding[norm(o["name"])] = int(m.group(0))

    with open(os.path.join(OUT, "leadership_history.csv"), newline="", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            if r["start_year"]:
                y = int(r["start_year"])
                k = norm(r["org_name"])
                founding[k] = min(y, founding.get(k, y))
    return founding


def build():
    with open(os.path.join(DATA, "ed_timeline_sample.txt"), encoding="utf-8") as f:
        sample = [l.strip() for l in f if l.strip()]

    lead_by_org = load_leadership_history()
    founding_by_org = load_founding_years()

    with open(os.path.join(WEB_DATA, "dataset.json"), encoding="utf-8") as f:
        dataset = json.load(f)
    ds_by_org = {norm(o["name"]): o for o in dataset["orgs"]}

    CURRENT_YEAR = 2025

    rows = []  # (org_name, {year: cell_text})
    known_founding = []
    for name in sample:
        k = norm(name)
        entries = lead_by_org.get(k, [])
        ds = ds_by_org.get(k)
        founding = founding_by_org.get(k) or (ds["firstYear"] if ds and ds.get("firstYear") else None)
        if founding:
            known_founding.append(founding)

        cells = {}
        undated_current = None  # a known current leader with no parseable start year
        for e in entries:
            if not e["start_year"]:
                # Some research entries name the current leader without a
                # parseable start ("current", "present (interim)") -- don't
                # silently drop them, just can't place them on the timeline
                # by year until filled in below.
                if e["ongoing"] == "yes":
                    undated_current = e["person"]
                continue
            start = int(e["start_year"])
            end = int(e["end_year"]) if e["end_year"] else (CURRENT_YEAR if e["ongoing"] == "yes" else start)
            full, ini = initials(e["person"])
            for y in range(start, min(end, CURRENT_YEAR) + 1):
                cells[y] = f"{full} ({ini})" if y == start else ini

        last_covered = max(cells.keys()) if cells else None
        if last_covered is None or last_covered < CURRENT_YEAR:
            # Prefer a named-but-undated current leader from direct research
            # over the pipeline's own EIN-matched executive when both exist,
            # since the former usually carries a title/context the latter
            # doesn't; fall back to the pipeline match otherwise.
            ex_name = undated_current or (
                ds["executive"]["name"] if ds and ds.get("executive") and ds["executive"].get("name") else None
            )
            if ex_name:
                fill_start = (last_covered + 1) if last_covered else (founding or CURRENT_YEAR)
                _, ini = initials(ex_name)
                for i, y in enumerate(range(fill_start, CURRENT_YEAR + 1)):
                    cells[y] = f"{ex_name} ({ini}) [current, exact start year unconfirmed]" if i == 0 else ini

        rows.append((name, founding, cells))

    start_year = min(known_founding) if known_founding else 1980
    years = list(range(start_year, CURRENT_YEAR + 1))

    wb = load_workbook(os.path.join(OUT, "corridor_dataset.xlsx"))
    if "ED Timeline (sample)" in wb.sheetnames:
        del wb["ED Timeline (sample)"]
    ws = wb.create_sheet("ED Timeline (sample)")

    navy = PatternFill("solid", fgColor="15294B")
    white = Font(color="FFFFFF", bold=True)
    na_fill = PatternFill("solid", fgColor="EDEDED")
    start_fill = PatternFill("solid", fgColor="EAF1FE")

    header = ["Organization"] + [str(y) for y in years]
    ws.append(header)
    for c in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = navy
        cell.font = white
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "B2"

    for name, founding, cells in rows:
        row_vals = [name]
        for y in years:
            if founding and y < founding:
                row_vals.append("N/A")
            else:
                row_vals.append(cells.get(y, ""))
        ws.append(row_vals)
        r = ws.max_row
        for ci, y in enumerate(years, start=2):
            val = ws.cell(row=r, column=ci).value
            if val == "N/A":
                ws.cell(row=r, column=ci).fill = na_fill
            elif val and "(" in val:
                ws.cell(row=r, column=ci).fill = start_fill

    ws.column_dimensions["A"].width = 46
    for c in range(2, len(years) + 2):
        ws.column_dimensions[get_column_letter(c)].width = 10

    note_row = ["NOTE: sample of 20 organizations drawn at random from the full roster (data/ed_timeline_sample.txt). "
                 "Years come primarily from direct research; where that data stops before the present and the "
                 "automated pipeline found a different named current executive, that name is added with its start "
                 "year marked unconfirmed. A blank cell means the organization existed but no leader is confirmed "
                 "for that year, not that the seat was vacant."]
    ws.append(note_row)

    path = os.path.join(OUT, "corridor_dataset.xlsx")
    wb.save(path)
    print(f"ED Timeline (sample): {len(rows)} orgs, {start_year}-{CURRENT_YEAR} ({len(years)} year columns)")


if __name__ == "__main__":
    build()
