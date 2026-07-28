"""
Single source of truth for both deliverables. Reads every CSV the scraper and
enricher produce and writes, from the same data:

  - web/src/data/dataset.json   the web app's data
  - output/corridor_dataset.xlsx  the formatted research spreadsheet
  - web/public/corridor_dataset.xlsx  + web/public/data/*.csv  download copies

Because the app and the spreadsheet are built here together, they never drift.

Run order:  corridor.py  ->  tools/enrich.py  ->  tools/build_outputs.py
"""

import csv
import json
import os
import re
import shutil
import unicodedata
from datetime import datetime, timezone
from difflib import SequenceMatcher

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(HERE, "output")
WEB_DATA = os.path.join(HERE, "web", "src", "data")
WEB_PUBLIC = os.path.join(HERE, "web", "public")

COMP_RE = re.compile(r"[^0-9]")

# Tax-credit report years whose source PDFs are hosted for download.
TAXCREDIT_YEARS = ("2015", "2016", "2017", "2018", "2019", "2020")


def norm_name(s):
    """Normalize an org name for cross-source matching."""
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode().lower()
    s = re.sub(r"\([^)]*\)", " ", s)
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    for w in ["community", "development", "corporation", "corp", "cdc", "inc",
              "incorporated", "association", "the", "of", "company", "companies",
              "group", "services", "center", "aka", "npi", "hfi"]:
        s = re.sub(rf"\b{w}\b", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def best_match(target, candidates):
    """Return (key, score) of the candidate name most similar to target."""
    tn = norm_name(target)
    best, score = None, 0.0
    for c in candidates:
        s = SequenceMatcher(None, tn, norm_name(c)).ratio()
        if s > score:
            best, score = c, s
    return best, score


def num(v):
    if v in (None, "", "None"):
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def money_int(s):
    if not s:
        return None
    d = COMP_RE.sub("", s)
    return int(d) if d else None


def load_csv(name):
    path = os.path.join(OUT, name)
    if not os.path.exists(path):
        return []
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def build():
    roster = load_csv("roster.csv")
    history = load_csv("financial_history.csv")
    leadership = load_csv("leadership.csv")
    news = load_csv("news.csv")
    taxcredit = load_csv("taxcredit.csv")
    filings = load_csv("filings.csv")
    geocoded = load_csv("geocoded.csv")
    geo_by_name = {g["org_name"]: g for g in geocoded}

    by_ein_hist, by_ein_lead, by_ein_news, by_ein_filings = {}, {}, {}, {}
    by_name_news = {}  # for orgs with no EIN (blank ein would collide across orgs)
    for h in history:
        by_ein_hist.setdefault(h["ein"], []).append(h)
    for l in leadership:
        by_ein_lead.setdefault(l["ein"], []).append(l)
    for n in news:
        if n["ein"]:
            by_ein_news.setdefault(n["ein"], []).append(n)
        else:
            by_name_news.setdefault(norm_name(n["org_name"]), []).append(n)
    for f in filings:
        by_ein_filings.setdefault(f["ein"], []).append(f)

    # Collapse the tax-credit reports into one record per participating org.
    tc_by_name = {}
    for t in taxcredit:
        rec = tc_by_name.setdefault(t["org_name"], {
            "name": t["org_name"], "startYear": num(t["start_year"]),
            "amount": None, "reportYears": []})
        rec["reportYears"].append(int(t["report_year"]))
        rec["amount"] = num(t["amount"]) or rec["amount"]
    for rec in tc_by_name.values():
        rec["reportYears"] = sorted(set(rec["reportYears"]))
    tc_latest = max((int(y) for t in taxcredit for y in [t["report_year"]]), default=None)

    orgs = []
    for r in roster:
        ein = r.get("ein", "")
        gaps = []
        for g in (r.get("filing_gaps") or "").split(";"):
            g = g.strip()
            if "->" in g:
                a, b = g.split("->")
                gaps.append({"from": int(a), "to": int(b)})
        comp_jumps = [int(x.strip()) for x in (r.get("comp_jump_years") or "").split(";") if x.strip().isdigit()]

        hist = []
        for h in sorted(by_ein_hist.get(ein, []), key=lambda x: num(x["year"]) or 0):
            hist.append({"year": num(h["year"]), "revenue": num(h["revenue"]),
                         "expenses": num(h["expenses"]), "assets": num(h["assets_eoy"]),
                         "employees": num(h["employees"]), "officerComp": num(h["officer_comp"])})

        officers = [{"name": o["name"], "title": o["title"], "comp": money_int(o["comp"]),
                     "isExecutive": o.get("is_executive") == "yes"}
                    for o in by_ein_lead.get(ein, [])]
        executive = next((o for o in officers if o["isExecutive"]), None)

        org_news = by_ein_news.get(ein, []) if ein else by_name_news.get(norm_name(r.get("org_name", "")), [])
        arts = sorted(org_news, key=lambda a: a.get("date", ""), reverse=True)
        articles = [{"title": a["title"], "source": a["source"], "url": a["url"], "date": a["date"]}
                    for a in arts]

        # Actual Form 990 PDFs (newest first).
        org_filings = [{"year": num(f["year"]), "pdfUrl": f["pdf_url"]}
                       for f in sorted(by_ein_filings.get(ein, []),
                                       key=lambda x: num(x["year"]) or 0, reverse=True)]

        # CDC Tax Credit participation, matched by name (works even without an EIN).
        tax_credit = None
        if tc_by_name:
            cand, score = best_match(r.get("org_name", ""), tc_by_name.keys())
            if cand and score >= 0.6:
                rec = tc_by_name[cand]
                tax_credit = {
                    "matchedName": rec["name"], "startYear": rec["startYear"],
                    "annualAmount": rec["amount"], "reportYears": rec["reportYears"],
                    "active": tc_latest in rec["reportYears"] if tc_latest else False,
                }

        last_year = num(r.get("last_year"))
        closed = bool(last_year and last_year <= 2021 and ein)
        instability = len(gaps) * 2 + len(comp_jumps) + (3 if closed else 0)

        geo = geo_by_name.get(r.get("org_name", ""))
        lat = float(geo["lat"]) if geo and geo.get("lat") else None
        lon = float(geo["lon"]) if geo and geo.get("lon") else None

        orgs.append({
            "name": r.get("org_name", ""), "type": r.get("type", ""), "ein": ein,
            "irsName": r.get("irs_name", ""), "irsCity": r.get("irs_city", ""),
            "confidence": r.get("match_confidence", ""),
            "firstYear": num(r.get("first_year")), "lastYear": last_year,
            "yearsFiled": num(r.get("years_filed")) or 0,
            "filingGaps": gaps, "compJumpYears": comp_jumps,
            "latestRevenue": num(r.get("latest_revenue")), "latestEmployees": num(r.get("latest_employees")),
            "website": r.get("website", ""), "contactEmail": r.get("contact_email", ""),
            "source": r.get("source", ""), "notes": r.get("notes", ""),
            "archiveCollection": r.get("archive_collection", ""), "address": r.get("address", ""),
            "lat": lat, "lon": lon,
            "closedCandidate": closed, "instability": instability,
            "executive": executive, "officers": officers, "leadershipAsOf": last_year,
            "news": articles, "filings": org_filings, "taxCredit": tax_credit,
            "history": hist,
        })

    orgs.sort(key=lambda o: o["name"].lower())
    matched = [o for o in orgs if o["ein"]]
    summary = {
        "total": len(orgs), "matched": len(matched),
        "withFinancialData": len([o for o in matched if o["history"]]),
        "highConfidence": len([o for o in matched if o["confidence"] == "high"]),
        "needsLookup": len([o for o in orgs if not o["ein"]]),
        "withSignal": len([o for o in orgs if o["filingGaps"] or o["compJumpYears"]]),
        "withNamedExecutive": len([o for o in orgs if o["executive"]]),
        "newsArticles": sum(len(o["news"]) for o in orgs),
        "inTaxCreditProgram": len([o for o in orgs if o["taxCredit"] and o["taxCredit"]["active"]]),
        "everTaxCredit": len([o for o in orgs if o["taxCredit"]]),
        "form990Pdfs": sum(len(o["filings"]) for o in orgs),
        "cdcs": len([o for o in orgs if o["type"] == "CDC"]),
        "bids": len([o for o in orgs if o["type"] == "BID"]),
        "closureCandidates": len([o for o in orgs if o["closedCandidate"]]),
        "mapped": len([o for o in orgs if o["lat"] and o["lon"]]),
        "earliestYear": min((o["firstYear"] for o in matched if o["firstYear"]), default=None),
        "latestYear": max((o["lastYear"] for o in matched if o["lastYear"]), default=None),
    }
    payload = {"generatedAt": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
               "summary": summary, "orgs": orgs}

    os.makedirs(WEB_DATA, exist_ok=True)
    with open(os.path.join(WEB_DATA, "dataset.json"), "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2)
    print(f"dataset.json: {summary['total']} orgs, {summary['withNamedExecutive']} with named exec, "
          f"{summary['everTaxCredit']} in tax-credit program, {summary['form990Pdfs']} 990 PDFs, "
          f"{summary['newsArticles']} news articles")

    write_xlsx(orgs, leadership, history, news, taxcredit, filings)
    copy_downloads()


def write_xlsx(orgs, leadership, history, news, taxcredit, filings):
    wb = Workbook()
    navy = PatternFill("solid", fgColor="15294B")
    white = Font(color="FFFFFF", bold=True)
    sig = PatternFill("solid", fgColor="FCEFC7")
    exec_fill = PatternFill("solid", fgColor="EAF1FE")
    tc_fill = PatternFill("solid", fgColor="E7F3EA")
    low_fill = PatternFill("solid", fgColor="F6D6D6")

    def sheet(ws, headers, rows, shade=None):
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = navy; cell.font = white
            cell.alignment = Alignment(vertical="center")
        for row in rows:
            ws.append(row)
            if shade:
                shade(ws, ws.max_row, row)
        ws.freeze_panes = "A2"
        for col in ws.columns:
            ln = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(ln + 2, 52)

    # Roster (with executive director + tax-credit participation)
    ws = wb.active; ws.title = "Roster"
    rrows = []
    for o in orgs:
        ex = o["executive"]
        tc = o["taxCredit"]
        rrows.append([o["name"], o["ein"],
                      ex["name"] if ex else "", ex["title"] if ex else "",
                      o["firstYear"], o["lastYear"], o["yearsFiled"],
                      o["latestRevenue"], o["latestEmployees"],
                      tc["startYear"] if tc else "", tc["annualAmount"] if tc else "",
                      "yes" if tc and tc["active"] else "",
                      "; ".join(f'{g["from"]}-{g["to"]}' for g in o["filingGaps"]),
                      "; ".join(str(y) for y in o["compJumpYears"]),
                      o["website"], o["contactEmail"], o["address"], o["source"], o["notes"]])
    def shade_roster(ws, r, row):
        if row[11] == "yes":
            ws.cell(row=r, column=12).fill = tc_fill
        if row[12] or row[13]:
            for c in (13, 14):
                ws.cell(row=r, column=c).fill = sig
    sheet(ws, ["Organization", "EIN", "Executive director", "Exec title",
               "First year", "Last year", "Years filed", "Latest revenue", "Latest staff",
               "Tax credit since", "Annual amount", "Tax credit active",
               "Filing gaps", "Pay-shift years", "Website", "Contact", "Address",
               "Roster source", "Notes"], rrows, shade_roster)

    # Temple Archives finds (Ben's 2026-07 batch): the CDCs identified from
    # Temple University Special Collections Research Center finding aids that
    # weren't already in the roster, with what we could and couldn't confirm.
    wsta = wb.create_sheet("Temple Archives")
    archive_orgs = [o for o in orgs if o["source"] == "Temple archives"]
    tarows = []
    for o in archive_orgs:
        ex = o["executive"]
        inquirer_count = sum(1 for n in o["news"] if "inquirer" in (n["source"] or "").lower())
        tarows.append([
            o["name"], o["archiveCollection"], o["ein"] or "",
            "needs manual lookup" if not o["ein"] else o["confidence"],
            o["yearsFiled"], ex["name"] if ex else "",
            len(o["news"]), inquirer_count, o["notes"],
        ])
    def shade_temple(ws, r, row):
        if row[3] == "needs manual lookup":
            ws.cell(row=r, column=4).fill = low_fill
        if row[7]:
            ws.cell(row=r, column=8).fill = tc_fill
    sheet(wsta, ["Organization", "Archive collection", "EIN", "Match status",
                 "Years filed", "Executive director", "News articles found",
                 "Inquirer articles", "Notes"], tarows, shade_temple)
    wsta.column_dimensions["I"].width = 70

    # Leadership (full current rosters)
    ws2 = wb.create_sheet("Leadership")
    lrows = [[l["org_name"], l["ein"], l["name"], l["title"], money_int(l["comp"]),
              "yes" if l.get("is_executive") == "yes" else ""] for l in leadership]
    def shade_lead(ws, r, row):
        if row[5] == "yes":
            for c in range(1, 7):
                ws.cell(row=r, column=c).fill = exec_fill
    sheet(ws2, ["Organization", "EIN", "Name", "Title", "Compensation", "Executive"], lrows, shade_lead)

    # Financial History
    ws3 = wb.create_sheet("Financial History")
    frows = [[h["org_name"], h["ein"], num(h["year"]), num(h["revenue"]), num(h["expenses"]),
              num(h["assets_eoy"]), num(h["employees"]), num(h["officer_comp"])] for h in history]
    sheet(ws3, ["Organization", "EIN", "Year", "Revenue", "Expenses", "Assets", "Staff", "Officer pay"], frows)

    # CDC Tax Credit participation (City of Philadelphia program)
    wstc = wb.create_sheet("Tax Credit")
    tcrows = [[t["org_name"], num(t["report_year"]), num(t["start_year"]), num(t["amount"])]
              for t in sorted(taxcredit, key=lambda x: (x["report_year"], x["org_name"].lower()))]
    sheet(wstc, ["Organization", "Report year", "In program since", "Annual contribution"], tcrows)

    # Actual Form 990 PDFs (one row per filing, links straight to the document)
    wsf = wb.create_sheet("990 PDFs")
    frrows = [[f["org_name"], num(f["year"]), f["pdf_url"]]
              for f in sorted(filings, key=lambda x: (x["org_name"].lower(), -(num(x["year"]) or 0)))]
    sheet(wsf, ["Organization", "Year", "Form 990 PDF"], frrows)

    # News & Sources
    ws4 = wb.create_sheet("News & Sources")
    nrows = [[n["org_name"], n["date"], n["source"], n["title"], n["url"]] for n in news]
    sheet(ws4, ["Organization", "Date", "Source", "Headline", "Link"], nrows)

    # Signals (ranked)
    ws5 = wb.create_sheet("Signals")
    ranked = sorted([o for o in orgs if o["ein"] and o["instability"] > 0],
                    key=lambda o: o["instability"], reverse=True)
    srows = [[o["name"], o["type"], (o["executive"] or {}).get("name", ""),
              len(o["compJumpYears"]), len(o["filingGaps"]),
              "yes" if o["closedCandidate"] else "", o["instability"]] for o in ranked]
    sheet(ws5, ["Organization", "Type", "Executive director", "Pay shifts", "Filing gaps",
                "Closure candidate", "Instability score"], srows)

    # About
    ws6 = wb.create_sheet("About")
    for line in ABOUT_LINES:
        ws6.append([line])
    ws6.column_dimensions["A"].width = 92
    ws6.cell(row=1, column=1).font = Font(bold=True, size=13)

    path = os.path.join(OUT, "corridor_dataset.xlsx")
    wb.save(path)
    print(f"corridor_dataset.xlsx: {len(wb.sheetnames)} sheets")


def copy_downloads():
    os.makedirs(os.path.join(WEB_PUBLIC, "data"), exist_ok=True)
    shutil.copy(os.path.join(OUT, "corridor_dataset.xlsx"),
                os.path.join(WEB_PUBLIC, "corridor_dataset.xlsx"))
    for f in ("roster.csv", "leadership.csv", "financial_history.csv", "news.csv",
              "taxcredit.csv", "filings.csv", "leadership_history.csv"):
        src = os.path.join(OUT, f)
        if os.path.exists(src):
            shutil.copy(src, os.path.join(WEB_PUBLIC, "data", f))
    # Tax-credit source PDFs, for offline reference.
    os.makedirs(os.path.join(WEB_PUBLIC, "taxcredit"), exist_ok=True)
    for yr in TAXCREDIT_YEARS:
        src = os.path.join(HERE, "taxcredit", f"{yr}.pdf")
        if os.path.exists(src):
            shutil.copy(src, os.path.join(WEB_PUBLIC, "taxcredit", f"cdc-tax-credit-{yr}.pdf"))
    print("copied spreadsheet + CSVs + tax-credit PDFs to web/public for download")


ABOUT_LINES = [
    "Corridor — Philadelphia CDC dataset",
    "github.com/AdamMirmina/corridor   ·   corridor.adammirmina.com",
    "",
    "WHAT EACH SHEET IS",
    "Roster: one row per CDC, with its current executive director, key metrics, address, tax-credit",
    "  status, roster source, and any research notes.",
    "Temple Archives: the 29 CDCs identified from Temple University Special Collections Research",
    "  Center finding aids (2026-07), with which specific collection each came from and what could",
    "  and couldn't be confirmed about it.",
    "ED Timeline (sample): a year-by-year executive-director grid for 20 randomly sampled",
    "  organizations, founding to present, in the format of a BID executive-director timeline",
    "  reference used for this project.",
    "Leadership: the full current officer and board roster per organization, from IRS Form 990 Part VII.",
    "  Titles carrying a '(To MM/YYYY)' note mark an officer who was on the way out — a recorded transition.",
    "Financial History: revenue, expenses, assets, staff, and officer pay per organization-year.",
    "Tax Credit: participation in the City of Philadelphia CDC Tax Credit program by report year,",
    "  with the year each organization entered and its annual contribution. Source PDFs: phila.gov.",
    "990 PDFs: a direct link to every Form 990 document on file, by organization and year.",
    "News & Sources: recent articles per organization, including a dedicated Philadelphia Inquirer",
    "  search alongside general Google News coverage and leadership-change coverage.",
    "Signals: organizations ranked by how much their record suggests leadership instability.",
    "",
    "WHERE EVERY PIECE OF DATA CAME FROM",
    "The roster itself: three separate rounds, each tagged in the Roster sheet's 'Roster source' column.",
    "  - 'Research roster' (70 orgs): the project's original hand-compiled list, plus follow-up",
    "    research — years active, full historical leadership with tenure ranges, mission statements,",
    "    addresses, and notes — folded into this roster's Notes/Address columns and into the",
    "    Leadership and ED Timeline sheets.",
    "  - 'City CDC program' (9 orgs): the City of Philadelphia's own list of CDC Tax Credit program",
    "    participants (phila.gov).",
    "  - 'Temple archives' (29 orgs): read directly from 13 PDF finding aids from Temple University's",
    "    Special Collections Research Center (library.temple.edu/scrc) — the Philadelphia Association",
    "    of Community Development Corporations' own member records (SCRC 241), plus standalone",
    "    collections for Germantown Settlement, Weccacoe Development Association, Olde Kensington",
    "    Redevelopment Corporation/Senior Housing Associates, West Philadelphia Corporation, Kensington",
    "    Action Now, and the Regional Council of Neighborhood Organizations.",
    "  - 'PACDC member list' (1 org) and 'Regional Foundation partners page' (1 org): cross-referenced",
    "    PACDC's own FY25 Annual Report (pacdc.org) and the Regional Foundation's partners page",
    "    (regionalfoundation.org) against the existing roster to find CDCs not yet tracked.",
    "  - Two more ('Research roster'): Urban Resources Development Corporation and Eastwick United CDC,",
    "    surfaced from further research and independently verified before adding.",
    "  - Checked but yielded no new organizations: shelterforce.org (national context, no Philadelphia-",
    "    specific CDC names).",
    "Financials, EIN matching, current officers, and Form 990 links: the ProPublica Nonprofit Explorer",
    "  API (projects.propublica.org/nonprofits), matched to each roster name by fuzzy string match,",
    "  scored and flagged by confidence rather than guessed. Historical (year-by-year) director names",
    "  are NOT available this way — see the limitation note below.",
    "News coverage: Google News search per organization, plus a Philadelphia-Inquirer-specific",
    "  site-scoped search so real Inquirer coverage isn't crowded out by more numerous generic sources.",
    "Tax-credit participation: the City of Philadelphia's own annual CDC Tax Credit program reports",
    "  (phila.gov), parsed from the source PDFs (2015, 2019, 2020 and other available years).",
    "",
    "THE HONEST LIMITATION",
    "Executive directors and boards from ProPublica are current as of each organization's latest Form",
    "  990, not a full year-by-year history — as of 2026 there is no free, lightweight way to pull",
    "  historical Form 990 Part VII names at scale (the AWS 990 mirror was retired, the IRS per-file",
    "  XML URLs 404, and ProPublica blocks scripted PDF downloads). Where a fuller history exists, it's",
    "  because direct research found it by hand — that's what powers the Leadership tenure ranges",
    "  in the Roster notes and the ED Timeline sheet, not an automated source. The pay-shift years flag",
    "  when a change most likely happened from officer-compensation swings; the news links are where",
    "  the actual named change is reported.",
]


if __name__ == "__main__":
    build()
