"""
Corridor — Philadelphia CDC/BID leadership + history tracker.

Builds the foundational dataset for the research question:
"What factors determine leadership transitions in community economic
development organizations in Philadelphia?"

What it does, per organization in the roster:
  1. Matches the org to its IRS EIN via the ProPublica Nonprofit Explorer API.
  2. Pulls every year of digitized Form 990 data: revenue, assets, employee
     count (structural size), and the years it filed (operational lifespan).
  3. Derives turnover SIGNALS from that history: filing gaps, the IRS
     "final return" flag, and year-over-year jumps in officer compensation
     (a large comp change is often the fingerprint of an ED change).

What it deliberately does NOT do yet: pull the executive director's NAME per
year. Those names live in Part VII of each 990, and as of 2026 every free
lightweight path to them is gone (the AWS 990 mirror was decommissioned, the
IRS per-file XML URLs 404, ProPublica blocks scripted PDF pulls). The only
remaining free route is the IRS bulk XML ZIPs, which is a heavier second stage.
So the Leadership sheet is generated as a structured fill-in template, keyed to
the years and the comp signals, so the manual lookup is fast and targeted.

Sources, all free, no API key:
  - ProPublica Nonprofit Explorer API  (financials, EIN, filing years)
  - PACDC member list                  (CDC roster)
  - OpenDataPhilly Business Improvement Districts  (BID roster + contacts)

Run:  py corridor.py
Out:  output/corridor_dataset.xlsx  and  output/*.csv
"""

import csv
import json
import os
import re
import sys
import time
import urllib.parse
from difflib import SequenceMatcher

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(HERE, "data")
OUT = os.path.join(HERE, "output")
PP = "https://projects.propublica.org/nonprofits/api/v2"
HEADERS = {"User-Agent": "corridor (academic research; github.com/AdamMirmina/corridor)"}

# A comp change at least this large year-over-year is flagged as a possible
# leadership transition. Tuned conservatively; it is a lead, not a conclusion.
COMP_JUMP_FRACTION = 0.30


def norm(name: str) -> str:
    """Normalize an org name for fuzzy comparison."""
    name = name.lower()
    name = re.sub(r"[^a-z0-9 ]", " ", name)
    # Drop boilerplate words that add noise to the match.
    for w in ["the", "inc", "incorporated", "corporation", "corp", "cdc",
              "community", "development", "association", "association",
              "company", "companies", "group", "of", "and"]:
        name = re.sub(rf"\b{w}\b", " ", name)
    return re.sub(r"\s+", " ", name).strip()


def similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, norm(a), norm(b)).ratio()


def load_roster() -> list[dict]:
    # Scope is CDCs only (the BID roster in data/bid_roster.csv is retained for
    # reference but no longer loaded).
    rows = []
    with open(os.path.join(DATA, "cdc_roster.csv"), newline="", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            rows.append({"org_name": r["org_name"], "type": r["type"],
                         "website": "", "contact_email": "", "source": r["source"],
                         "notes": r.get("notes", ""),
                         "archive_collection": r.get("archive_collection", ""),
                         "address": r.get("address", "")})
    seen, deduped = set(), []
    for r in rows:
        key = norm(r["org_name"])
        if key in seen:
            continue
        seen.add(key)
        deduped.append(r)
    return deduped


ABBREV = {
    "cdc": "community development corporation",
    "bid": "business improvement district",
    "nac": "neighborhood advisory committee",
    "clt": "community land trust",
}


def query_variants(org_name: str) -> list[str]:
    """ProPublica search does literal token matching, so 'Nicetown CDC' finds
    nothing while 'Nicetown Community Development Corporation' finds it. Build a
    few progressively looser queries and try each until one returns results."""
    import unicodedata

    variants = []
    # Strip accents and parenthetical acronyms (e.g. "(APM)", "(LA21)"), and
    # split CamelCase the roster sometimes uses ("ProjectHOME", "NewCourtland").
    base = unicodedata.normalize("NFKD", org_name).encode("ascii", "ignore").decode()
    base = re.sub(r"\([^)]*\)", " ", base)
    base = re.sub(r"(?<=[a-z])(?=[A-Z])", " ", base)
    base = re.sub(r"\s+", " ", base).strip()

    def add(v):
        v = re.sub(r"\s+", " ", v).strip()
        if v and v.lower() not in [x.lower() for x in variants]:
            variants.append(v)

    add(org_name)  # the literal name first
    add(base)
    # Expand known abbreviations (CDC -> Community Development Corporation).
    expanded = base
    for ab, full in ABBREV.items():
        expanded = re.sub(rf"\b{ab}\b", full, expanded, flags=re.IGNORECASE)
    expanded = expanded.replace("/", " ")
    add(expanded)
    # Drop the org-type tokens entirely and search the distinctive core.
    core = re.sub(r"\b(cdc|bid|nac|clt|inc|incorporated|corp|corporation|"
                  r"company|companies|district|the)\b", " ", base, flags=re.IGNORECASE)
    core = core.replace("/", " ")
    add(core)
    # Last resort: first two significant words of the core.
    words = [w for w in core.split() if len(w) > 2]
    if len(words) >= 2:
        add(" ".join(words[:2]))
    return variants


def search_ein(org_name: str) -> dict:
    """Find the best-matching PA/Philadelphia EIN for an org name, trying
    several query variants and scoring every candidate against the real name."""
    candidates = {}  # ein -> org dict (first seen wins)
    used_query = ""
    for q in query_variants(org_name):
        url = f"{PP}/search.json?" + urllib.parse.urlencode({"q": q, "state[id]": "PA"})
        try:
            data = requests.get(url, headers=HEADERS, timeout=30).json()
        except Exception as e:
            return {"ein": "", "matched_name": "", "city": "", "confidence": "error",
                    "score": 0.0, "note": str(e)[:80]}
        orgs = data.get("organizations", [])
        if orgs and not used_query:
            used_query = q
        for o in orgs:
            candidates.setdefault(o["ein"], o)
        time.sleep(0.1)
        if len(candidates) >= 25:
            break

    if not candidates:
        return {"ein": "", "matched_name": "", "city": "", "confidence": "none",
                "score": 0.0, "note": "no PA results", "query": ""}

    best, best_rank = None, -1.0
    for o in candidates.values():
        s = similarity(org_name, o.get("name", ""))
        rank = s + (0.08 if (o.get("city") or "").lower() == "philadelphia" else 0)
        if rank > best_rank:
            best, best_rank = o, rank
    base = similarity(org_name, best.get("name", ""))
    # Reject weak matches outright. A wrong EIN silently poisons the dataset;
    # an honest blank sends the org to the manual-lookup pile instead. Acronym
    # orgs (SEAMAAC, HACE) land here because their registered name is the
    # spelled-out version, which no fuzzy match against the acronym will catch.
    if base < 0.45:
        return {"ein": "", "matched_name": "", "city": "", "confidence": "none",
                "score": round(base, 2), "query": used_query,
                "note": f"rejected weak match: {best.get('name','')[:40]}"}
    conf = "high" if base >= 0.72 else "medium" if base >= 0.5 else "low"
    # ProPublica's API returns some EINs as bare JSON integers rather than
    # zero-padded strings, which silently drops a leading zero (a real EIN is
    # always 9 digits). Zero-pad to restore it.
    return {"ein": str(best["ein"]).zfill(9), "matched_name": best.get("name", ""),
            "city": best.get("city", ""), "confidence": conf,
            "score": round(base, 2), "note": "", "query": used_query}


def fetch_history(ein: str) -> dict:
    """Pull all digitized 990 years for an EIN and derive size + lifespan."""
    url = f"{PP}/organizations/{ein}.json"
    try:
        data = requests.get(url, headers=HEADERS, timeout=30).json()
    except Exception as e:
        return {"error": str(e)[:80], "filings": []}
    org = data.get("organization", {})
    filings = data.get("filings_with_data", [])
    rows = []
    for f in filings:
        yr = f.get("tax_prd_yr")
        if yr is None:
            continue
        rows.append({
            "year": yr,
            "revenue": f.get("totrevenue"),
            "expenses": f.get("totfuncexpns"),
            "assets_eoy": f.get("totassetsend"),
            "net_assets_eoy": f.get("totnetassetend"),
            "employees": f.get("totemployee"),
            "volunteers": f.get("totvolunteers"),
            "officer_comp": f.get("compnsatncurrofcr"),
        })
    rows.sort(key=lambda r: r["year"])
    return {"error": "", "org_meta": org, "rows": rows}


def analyze(rows: list[dict]) -> dict:
    """Compute lifespan + turnover signals from the filing history."""
    if not rows:
        return {"first_year": "", "last_year": "", "years_filed": 0,
                "filing_gaps": "", "comp_jump_years": "", "latest_revenue": "",
                "latest_employees": ""}
    years = [r["year"] for r in rows]
    gaps = []
    for a, b in zip(years, years[1:]):
        if b - a > 1:
            gaps.append(f"{a}->{b}")
    # Year-over-year officer comp jumps (possible ED change).
    jump_years = []
    prev = None
    for r in rows:
        c = r["officer_comp"]
        if prev is not None and prev > 0 and c is not None:
            change = abs(c - prev) / prev
            if change >= COMP_JUMP_FRACTION:
                jump_years.append(r["year"])
        if c is not None and c > 0:
            prev = c
    latest = rows[-1]
    return {
        "first_year": years[0],
        "last_year": years[-1],
        "years_filed": len(years),
        "filing_gaps": "; ".join(gaps),
        "comp_jump_years": "; ".join(str(y) for y in jump_years),
        "latest_revenue": latest["revenue"],
        "latest_employees": latest["employees"],
    }


def run():
    os.makedirs(OUT, exist_ok=True)
    roster = load_roster()
    print(f"Roster: {len(roster)} organizations")

    enriched = []
    history_long = []   # one row per org-year
    leadership_rows = []  # fill-in template

    for i, org in enumerate(roster, 1):
        name = org["org_name"]
        print(f"[{i:>2}/{len(roster)}] {name[:48]:48} ", end="", flush=True)
        m = search_ein(name)
        rec = dict(org)
        rec.update({"ein": m["ein"], "irs_name": m["matched_name"],
                    "irs_city": m["city"], "match_confidence": m["confidence"],
                    "match_score": m["score"]})
        if m["ein"]:
            hist = fetch_history(m["ein"])
            rows = hist.get("rows", [])
            a = analyze(rows)
            rec.update(a)
            for r in rows:
                history_long.append({"org_name": name, "ein": m["ein"], **r})
                leadership_rows.append({
                    "org_name": name, "ein": m["ein"], "year": r["year"],
                    "officer_comp": r["officer_comp"],
                    "executive_director": "",  # manual fill
                    "title": "", "source_url": "", "notes": ""})
            print(f"EIN {m['ein']}  {m['confidence']:6}  {a['years_filed']:>2}y  "
                  f"gaps[{a['filing_gaps']}]  comp-jumps[{a['comp_jump_years']}]")
        else:
            rec.update({"first_year": "", "last_year": "", "years_filed": 0,
                        "filing_gaps": "", "comp_jump_years": "",
                        "latest_revenue": "", "latest_employees": ""})
            print(f"NO MATCH ({m['note']})")
        enriched.append(rec)
        time.sleep(0.25)  # be polite to ProPublica

    write_csvs(enriched, history_long, leadership_rows)
    write_xlsx(enriched, history_long, leadership_rows)
    summarize(enriched)


def write_csvs(enriched, history_long, leadership_rows):
    def dump(path, rows, fields):
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=fields)
            w.writeheader()
            for r in rows:
                w.writerow({k: r.get(k, "") for k in fields})
    dump(os.path.join(OUT, "roster.csv"), enriched, list(enriched[0].keys()))
    if history_long:
        dump(os.path.join(OUT, "financial_history.csv"), history_long,
             list(history_long[0].keys()))
    if leadership_rows:
        dump(os.path.join(OUT, "leadership_template.csv"), leadership_rows,
             list(leadership_rows[0].keys()))


def write_xlsx(enriched, history_long, leadership_rows):
    wb = Workbook()
    hdr_fill = PatternFill("solid", fgColor="15294B")
    hdr_font = Font(color="FFFFFF", bold=True)
    low_fill = PatternFill("solid", fgColor="F6D6D6")  # low-confidence match
    sig_fill = PatternFill("solid", fgColor="FCEFC7")  # turnover signal

    def style_header(ws, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"

    def autowidth(ws, maxw=46):
        for col in ws.columns:
            ln = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(ln + 2, maxw)

    # Sheet 1: Roster (the Week-2 deliverable)
    ws = wb.active
    ws.title = "Roster"
    cols = ["org_name", "type", "ein", "irs_name", "irs_city", "match_confidence",
            "match_score", "first_year", "last_year", "years_filed", "filing_gaps",
            "comp_jump_years", "latest_revenue", "latest_employees",
            "website", "contact_email", "source"]
    ws.append([c.replace("_", " ").title() for c in cols])
    for r in enriched:
        ws.append([r.get(c, "") for c in cols])
        if r.get("match_confidence") == "low" or r.get("match_confidence") == "none":
            ws.cell(row=ws.max_row, column=6).fill = low_fill
        if r.get("comp_jump_years") or r.get("filing_gaps"):
            for c in (11, 12):
                ws.cell(row=ws.max_row, column=c).fill = sig_fill
    style_header(ws, len(cols))
    autowidth(ws)

    # Sheet 2: Financial History (size + lifespan, one row per org-year)
    ws2 = wb.create_sheet("Financial History")
    if history_long:
        cols2 = list(history_long[0].keys())
        ws2.append([c.replace("_", " ").title() for c in cols2])
        for r in history_long:
            ws2.append([r.get(c, "") for c in cols2])
        style_header(ws2, len(cols2))
        autowidth(ws2)

    # Sheet 3: Leadership template (manual fill, keyed to comp signals)
    ws3 = wb.create_sheet("Leadership (fill-in)")
    if leadership_rows:
        cols3 = list(leadership_rows[0].keys())
        ws3.append([c.replace("_", " ").title() for c in cols3])
        for r in leadership_rows:
            ws3.append([r.get(c, "") for c in cols3])
        style_header(ws3, len(cols3))
        autowidth(ws3)

    # Sheet 4: Notes
    ws4 = wb.create_sheet("How to read this")
    notes = [
        ["Corridor — Philadelphia CDC / BID leadership + history dataset"],
        ["Generated by Corridor (github.com/AdamMirmina/corridor)."],
        [""],
        ["Roster sheet"],
        ["  One row per organization. EIN, IRS-registered name, and city come from"],
        ["  the ProPublica Nonprofit Explorer (IRS data). Check rows shaded red in"],
        ["  the Match Confidence column: those are low-confidence EIN matches that"],
        ["  need a human to confirm or correct the EIN before trusting their data."],
        [""],
        ["  First/Last Year and Years Filed describe operational lifespan from 990"],
        ["  filings. Filing Gaps (shaded) can mean dormancy, a lapse, or a merger."],
        ["  A long gap ending with no recent filing is a strong 'organizational"],
        ["  death' candidate to investigate."],
        [""],
        ["  Comp Jump Years (shaded) flag years where total officer compensation"],
        ["  moved 30%+ versus the prior filing. That is a LEAD on a leadership"],
        ["  transition, not proof. Confirm against the actual 990 or the org's site."],
        [""],
        ["Financial History sheet"],
        ["  One row per organization-year: revenue, expenses, assets, employee"],
        ["  count, officer compensation. This is the structural-size series."],
        [""],
        ["Leadership (fill-in) sheet"],
        ["  Pre-built grid of every org-year, with a blank Executive Director"],
        ["  column. Director NAMES are not auto-filled: as of 2026 there is no"],
        ["  free lightweight API for Form 990 Part VII names. Fill these from the"],
        ["  org website, its archived staff pages (web.archive.org), and the 990"],
        ["  PDFs on ProPublica. Start with the Comp Jump years, they are where a"],
        ["  transition most likely happened."],
    ]
    for row in notes:
        ws4.append(row)
    ws4.column_dimensions["A"].width = 78
    ws4.cell(row=1, column=1).font = Font(bold=True, size=13)

    path = os.path.join(OUT, "corridor_dataset.xlsx")
    wb.save(path)
    print(f"\nWrote {path}")


def summarize(enriched):
    matched = [r for r in enriched if r.get("ein")]
    high = [r for r in matched if r.get("match_confidence") == "high"]
    low = [r for r in matched if r.get("match_confidence") in ("low", "none")]
    with_signals = [r for r in matched if r.get("comp_jump_years") or r.get("filing_gaps")]
    print("\n=== summary ===")
    print(f"  organizations:        {len(enriched)}")
    print(f"  matched to an EIN:    {len(matched)}")
    print(f"  high-confidence:      {len(high)}")
    print(f"  low/none (review):    {len(low)}")
    print(f"  with turnover signal: {len(with_signals)}")


if __name__ == "__main__":
    run()
